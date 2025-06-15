from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import random
import pandas as pd
from openpyxl import load_workbook

# ✅ Chrome の設定
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
chrome_options = Options()
chrome_options.add_argument(f"user-agent={user_agent}")
chrome_options.add_argument("--headless")  
chrome_options.add_argument("--disable-gpu")  
chrome_options.add_argument("--window-size=1920,1080")  # GPUエラー回避

# ChromeDriverのセットアップ
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

url = "https://salonboard.com/login/"

driver.get(url)
time.sleep(random.uniform(3,10))

#アカウント
user_name = "CB64315"
passwword = "anemone!1113"

#ID入力
name_element = WebDriverWait(driver,10).until(
EC.presence_of_element_located((By.CSS_SELECTOR,"div > dl:nth-child(1) > dd > input"))
)

if name_element:
    print("セレクタ見つけました")
else:
    print("セレクタが見つけられません")

name_element.clear()

name_element.send_keys(user_name)

time.sleep(random.uniform(5,7))

#パスワード入力
pass_element = WebDriverWait(driver,10).until(
EC.presence_of_element_located((By.CSS_SELECTOR,"#jsiPwInput"))
)

if pass_element:
    print("パスワードセレクタ発見")
else:
    print("セレクタ未発見")

pass_element.clear()

pass_element.send_keys(passwword)

time.sleep(random.uniform(4,8))

#enter
enter_element = WebDriverWait(driver,10).until(
EC.presence_of_element_located((By.CSS_SELECTOR,"#idPasswordInputForm > div > div > a"))
)

if enter_element:
       print("サロンボードにログインします…")
else:
    print("セレクタ未発見")
 
enter_element.click()


time.sleep(15)

#スクロールランダム
scroll_amount = random.randint(300, 1000)  # 300〜1000px の範囲でスクロール
driver.execute_script(f"window.scrollBy(0, {scroll_amount});")

shop_list = [
        "#H000222785","#H000230528","#H000232611","#H000239916","#H000254037",
        "#H000263792","#H000267119","#H000278723","#H000286626","#H000294209",
        "#H000304369","#H000311082","#H000313898","#H000331235","#H000334564",
        "#H000340031","#H000351618","#H000351620","#H000353753","#H000354418",
        "#H000356635","#H000377067","#H000379706","#H000401645","#H000413731",
        "#H000417538","#H000422770","#H000438662","#H000459606","#H000459605",
        "#H000467149","#H000467306","#H000480917","#H000494110","#H000502962",
        "#H000506434","#H000506436","#H000523761","#H000544821","#H000565670",
        "#H000574762","#H000592769","#H000596751","#H000637738","#H000641687",
        "#H000648370","#H000652001","#H000652003","#H000660884","#H000660887",
        "#H000664196","#H000664396","#H000664399","#H000669939","#H000674586",
        "#H000674594","#H000679801","#H000691442","#H000695065","#H000719432"
    ]

shop_data= []

#店舗ページ入る。
for shop_css in shop_list:
            
            shop_elements = WebDriverWait(driver,30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, shop_css))
        )
            shop_elements.click()

            print(f"店舗ページ {shop_css} に入りました。")

            # 300〜1000px の範囲でスクロール
            scroll_amount = random.randint(400, 1000) 
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")


            #予約管理
            header_element = driver.find_element(By.CSS_SELECTOR,"li:nth-child(1) > a")
            time.sleep(random.uniform(2,5))

            header_element.click()
            print("予約管理画面に入りました。")

        #予約一覧
            header_element_sec = WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"#globalNaviSub > div > div > ul > li:nth-child(2) > a")
        ))

            time.sleep(random.uniform(1,4))

            header_element_sec.click()
            print("予約一覧画面に入りました。")

        # 300〜1000px の範囲でスクロール
            scroll_amount = random.randint(400, 1200)  
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")

            time.sleep(20)


            #昨日CSSセレクター指定
            day_element = WebDriverWait(driver,10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,"#yesterdaySet"))
        )

            # 300〜1000px の範囲でスクロール
            scroll_amount = random.randint(400, 1200)  
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")


            #昨日を指定
            day_element.click()


            #済みCSS指定
            option_element = WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"#completed"))
        )

            #済み指定
            option_element.click()


            #会計済みCSS指定
            optional_element = WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"#salesRegistered"))
        )

            #会計済み指定
            optional_element.click()

            time.sleep(random.uniform (1,5))

        #検索CSS指定
            options_element = WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"#search"))
        )

        #検索済み指定
            options_element.click()
            if options_element:
                print("検索完了。読み込み中…")


            points = WebDriverWait(driver,10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR,"#reserveListArea tr td:nth-child(7) > b"))
            )
        

            shop_points = []

        

            points = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#reserveListArea tr td:nth-child(7) > b"))
            )


            for elem in points:
                    text = elem.text.strip()  # 余計なスペースを削除
            if text.isdigit():  # 数字かどうかチェック
                    shop_points.append(int(text))  # 数値に変換してリストに追加
            else:
                    print(f"数値に変換できない値: {text}")


            #２ページ目のセレクタ指定
            next_page = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,"#sortList > div.mod_column02.mt15.cf > div.columnBlock02 > div > p.next_disable > span"))
                    )
            

            if next_page.is_enabled():
                next_page.click()  # クリックして2ページ目へ移動

                # 2ページ目のデータを取得
                points = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#reserveListArea tr td:nth-child(7) > b"))
                )

                for elem in points:
                    text = elem.text.strip()  # 余計なスペースを削除
                    if text.isdigit():  # 数字かどうかチェック
                        shop_points.append(int(text))  # 数値に変換してリストに追加
                    else:
                        print(f"数値に変換できない値: {text}")

            else:
                    print("2ページ目が存在しません。処理をスキップします。")

    except Exception as e:

print(f"エラーが発生しました: {e}")            #２ページ目クリック

          

    
        
                # 合計を計算
                            
    total_points = sum(shop_points)

print(f"店舗 {shop_css} の合計ポイント: {total_points}")

shop_data.append([shop_css,total_points])

next_page = WebDriverWait(driver,10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,"#sortList > div.mod_column02.mt15.cf > div.columnBlock02 > div > p.next_disable > span"))
            )

        
            
            
back_element = WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,"#logo > a > img"))
            )

back_element.click()

time.sleep(random.uniform (2,7))

salon_listPage = WebDriverWait(driver,10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,"body > div.contents03 > div > div > div.columnBlock.w324 > div.mod_box_19.mb20 > div > a"))
            )

salon_listPage.click()

time.sleep(random.uniform (2,7))


            
file_path = r"C:\Users\littl\Downloads\SB_points.xlsx"

try:
    wb = load_workbook(file_path)
    ws = wb.active

    # ヘッダーを追加（最初の行が空なら）
    if ws.max_row == 1:
        ws.append(["店舗ID", "合計ポイント"])

    # 店舗データをExcelに追加
    for data in shop_data:
        ws.append(data)

    wb.save(file_path)
    wb.close()
    print("データをExcelに保存しました。")

except Exception as e:
    print(f"Excel保存エラー: {e}")

driver.quit()





